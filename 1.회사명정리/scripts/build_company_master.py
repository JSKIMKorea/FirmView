"""
DART 공식 법인명 매핑 (build_company_master.py)
===============================================
사용법:
  python build_company_master.py            # 매핑 실행
  python build_company_master.py --refresh  # DART 목록 강제 재다운로드

워크플로:
  1. 첫 실행  → ../output/review_needed.xlsx 생성 (검토 필요 항목)
  2. Excel 검토 후 ../company_overrides.json 에 보정값 추가
         또는 python import_review.py 실행으로 자동 반영
  3. 재실행   → 오버라이드 적용, review_needed.xlsx 항목 감소

출력물 (../output/):
  dart_corp_list.csv          DART 법인 목록 캐시 (7일 유효)
  company_master.json         고유 회사 목록 (웹 검색용)
  project_with_company.json   프로젝트 + 회사 통합 (최종 기초자료)
  review_needed.xlsx          수동 검토 필요 항목

설정 파일 (../):
  company_overrides.json      수동 보정 파일
"""

import sys, os, re, json, io, zipfile, requests
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
from rapidfuzz import process as fuzz_proc, fuzz

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────────────────────

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
BASE_DIR       = os.path.dirname(SCRIPT_DIR)          # 1.회사명정리/
OUTPUT_DIR     = os.path.join(BASE_DIR, "output")
PROJECT_JSON   = os.path.join(OUTPUT_DIR, "project_master.json")
CORP_CACHE     = os.path.join(OUTPUT_DIR, "dart_corp_list.csv")
OVERRIDES_FILE = os.path.join(BASE_DIR, "company_overrides.json")

# ─────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────

DART_API_KEY = "8d35c980d2f4947d4a345a5352ef3d1109133f7a"

FUZZY_AUTO = 92   # 이상 → 자동 확정
FUZZY_MIN  = 75   # 이상 → 검토 대상 / 미만 → 미발견

# ─────────────────────────────────────────────────────────────
# DART 법인 목록
# ─────────────────────────────────────────────────────────────

def _download_dart() -> pd.DataFrame:
    print("DART 법인 목록 다운로드 중 (최초 1회)...")
    url = f"https://opendart.fss.or.kr/api/corpCode.xml?crtfc_key={DART_API_KEY}"
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    xml_bytes = zf.read(zf.namelist()[0])
    root = ET.fromstring(xml_bytes)

    rows = []
    for item in root.findall(".//list"):
        code  = (item.findtext("corp_code")  or "").strip()
        name  = (item.findtext("corp_name")  or "").strip()
        stock = (item.findtext("stock_code") or "").strip()
        if code and name:
            rows.append({"corp_code": code, "corp_name": name, "stock_code": stock})
    df = pd.DataFrame(rows)
    print(f"  {len(df):,}개 법인 로드")
    return df


def load_dart(force=False) -> pd.DataFrame:
    if not force and os.path.exists(CORP_CACHE):
        age = (datetime.now().timestamp() - os.path.getmtime(CORP_CACHE)) / 86400
        if age < 7:
            print(f"  DART 캐시 사용 ({age:.1f}일 전, 7일 이내)")
            return pd.read_csv(CORP_CACHE, dtype=str).fillna("")
        print(f"  캐시 {age:.0f}일 경과 → 재다운로드")
    df = _download_dart()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df.to_csv(CORP_CACHE, index=False, encoding="utf-8-sig")
    return df

# ─────────────────────────────────────────────────────────────
# 정규화
# ─────────────────────────────────────────────────────────────

_LEGAL = re.compile(
    r"^\(주\)\s*|\s*\(주\)$"
    r"|^\(유\)\s*|\s*\(유\)$"
    r"|^\(사\)\s*|\s*\(사\)$"
    r"|^\(자\)\s*|\s*\(자\)$"
    r"|^주식회사\s*|\s*주식회사$"
    r"|^유한회사\s*|\s*유한회사$"
    r"|^합자회사\s*|\s*합자회사$"
    r"|^사단법인\s*|\s*사단법인$"
    r"|^재단법인\s*|\s*재단법인$"
    r"|^농업회사법인\s*"
)

def normalize(name: str) -> str:
    return _LEGAL.sub("", name.strip()).strip()


def extract_candidate(prjtnm: str) -> str:
    return prjtnm.split("/")[0].strip()

# ─────────────────────────────────────────────────────────────
# DART 인덱스 구축
# ─────────────────────────────────────────────────────────────

def build_index(dart_df: pd.DataFrame):
    dart_df = dart_df.reset_index(drop=True).copy()
    dart_df["norm_name"] = dart_df["corp_name"].apply(normalize)

    exact_idx, norm_idx = {}, {}
    for _, row in dart_df.iterrows():
        if row["corp_name"] not in exact_idx:
            exact_idx[row["corp_name"]] = row
        nn = row["norm_name"]
        if nn and nn not in norm_idx:
            norm_idx[nn] = row

    return dart_df, exact_idx, norm_idx

# ─────────────────────────────────────────────────────────────
# 매칭
# ─────────────────────────────────────────────────────────────

def _row_to_result(row, match_type, confidence) -> dict:
    return {
        "official_name": row["corp_name"],
        "corp_code":     row["corp_code"],
        "stock_code":    row["stock_code"],
        "match_type":    match_type,
        "confidence":    round(float(confidence), 3),
    }

_NOT_FOUND = {"official_name": "", "corp_code": "", "stock_code": "",
              "match_type": "not_found", "confidence": 0.0}


def match_one(candidate, dart_df, exact_idx, norm_idx, norm_names) -> dict:
    norm_cand = normalize(candidate)

    if candidate in exact_idx:
        return _row_to_result(exact_idx[candidate], "exact", 1.0)

    if norm_cand and norm_cand in norm_idx:
        return _row_to_result(norm_idx[norm_cand], "exact_normalized", 0.95)

    if norm_cand:
        hit = fuzz_proc.extractOne(
            norm_cand, norm_names,
            scorer=fuzz.token_set_ratio,
            score_cutoff=FUZZY_MIN,
        )
        if hit:
            _, score, idx = hit
            mtype = "fuzzy_auto" if score >= FUZZY_AUTO else "fuzzy_review"
            return _row_to_result(dart_df.iloc[idx], mtype, score / 100)

    return _NOT_FOUND.copy()


def get_alts(norm_cand, dart_df, norm_names, n=3) -> str:
    if not norm_cand:
        return ""
    tops = fuzz_proc.extract(norm_cand, norm_names,
                              scorer=fuzz.token_set_ratio, limit=n)
    return " | ".join(
        f"{dart_df.iloc[idx]['corp_name']}({score:.0f})"
        for _, score, idx in tops if score >= 50
    )

# ─────────────────────────────────────────────────────────────
# 오버라이드
# ─────────────────────────────────────────────────────────────

def load_overrides() -> dict:
    if not os.path.exists(OVERRIDES_FILE):
        return {}
    with open(OVERRIDES_FILE, encoding="utf-8") as f:
        return json.load(f).get("overrides", {})


def resolve_override(value, exact_idx, norm_idx) -> dict:
    if value == "__SKIP__":
        return {"official_name": "", "corp_code": "", "stock_code": "",
                "match_type": "__skip__", "confidence": 1.0}
    row = exact_idx.get(value) or norm_idx.get(normalize(value))
    if row is not None:
        return _row_to_result(row, "override", 1.0)
    return {"official_name": value, "corp_code": "", "stock_code": "",
            "match_type": "override_manual", "confidence": 1.0}

# ─────────────────────────────────────────────────────────────
# 저장
# ─────────────────────────────────────────────────────────────

def save_company_master(company_map) -> str:
    companies = [
        {
            "candidate_name": cand,
            "official_name":  info["official_name"] or cand,
            "corp_code":      info["corp_code"],
            "stock_code":     info["stock_code"],
            "match_type":     info["match_type"],
            "confidence":     info["confidence"],
        }
        for cand, info in sorted(company_map.items(),
                                  key=lambda x: x[1].get("official_name", x[0]))
        if info["match_type"] != "__skip__"
    ]
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total": len(companies),
        "companies": companies,
    }
    path = os.path.join(OUTPUT_DIR, "company_master.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  [OK] company_master.json        : {len(companies):,}개")
    return path


def save_review_excel(review_rows) -> str | None:
    if not review_rows:
        print("  검토 항목 없음 - review_needed.xlsx 생성 건너뜀")
        return None

    cols = [
        "candidate_name",
        "top_dart_match",
        "match_score(%)",
        "alt_matches (상위3)",
        "official_name_입력란",
        "corp_code_입력란 (선택)",
        "비고",
    ]
    df = pd.DataFrame(review_rows, columns=cols)
    path = os.path.join(OUTPUT_DIR, "review_needed.xlsx")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="검토필요", index=False)
        ws = writer.sheets["검토필요"]
        for col, w in zip(ws.columns, [30, 30, 14, 55, 30, 22, 20]):
            ws.column_dimensions[col[0].column_letter].width = w
        from openpyxl.styles import PatternFill, Font
        hdr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        inp_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.fill = inp_fill

    print(f"  [OK] review_needed.xlsx         : {len(df):,}건 검토 필요")
    return path


def save_project_with_company(projects, company_map) -> str:
    enriched = []
    for p in projects:
        cand = extract_candidate(p["name"])
        info = company_map.get(cand, _NOT_FOUND)
        enriched.append({
            "id":                p["id"],
            "name":              p["name"],
            "company_candidate": cand,
            "company_official":  info["official_name"] or cand,
            "corp_code":         info["corp_code"],
            "stock_code":        info["stock_code"],
            "match_type":        info["match_type"],
            "confidence":        info["confidence"],
        })
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total": len(enriched),
        "projects": enriched,
    }
    path = os.path.join(OUTPUT_DIR, "project_with_company.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  [OK] project_with_company.json  : {len(enriched):,}건")
    return path


def ensure_overrides_template():
    if os.path.exists(OVERRIDES_FILE):
        return
    template = {
        "_comment": [
            "candidate_name 을 키, 값에 DART 등록 공식 법인명 입력",
            "DART 미등록 외국계/소규모: 공식 상호 직접 입력",
            "회사명 아닌 항목: '__SKIP__' 입력",
            "재실행 시 이 파일이 자동 매칭보다 우선 적용됩니다",
        ],
        "overrides": {
            "예시_약칭": "정식법인명(주)",
            "예시_코드형": "__SKIP__",
        },
    }
    with open(OVERRIDES_FILE, "w", encoding="utf-8") as f:
        json.dump(template, f, ensure_ascii=False, indent=2)
    print(f"  [OK] company_overrides.json 템플릿 생성")

# ─────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────

def main():
    force_refresh = "--refresh" in sys.argv

    print("=" * 60)
    print("  DART 공식 법인명 매핑")
    print("=" * 60)

    dart_df = load_dart(force=force_refresh)
    dart_df, exact_idx, norm_idx = build_index(dart_df)
    norm_names = dart_df["norm_name"].tolist()

    with open(PROJECT_JSON, encoding="utf-8") as f:
        project_data = json.load(f)
    projects = project_data["projects"]
    print(f"\n프로젝트 {len(projects):,}건 로드")

    overrides = load_overrides()
    print(f"수동 오버라이드 {len(overrides)}건 로드")

    candidates = sorted({extract_candidate(p["name"]) for p in projects})
    print(f"고유 회사 후보: {len(candidates):,}개")

    print(f"\n매칭 중... (자동={FUZZY_AUTO} / 검토={FUZZY_MIN})")
    company_map, review_rows = {}, []
    stats = dict(exact=0, exact_normalized=0, fuzzy_auto=0,
                 fuzzy_review=0, override=0, not_found=0, skip=0)

    for i, cand in enumerate(candidates):
        if (i + 1) % 1000 == 0:
            print(f"  {i+1:,}/{len(candidates):,}...")

        if cand in overrides:
            result = resolve_override(overrides[cand], exact_idx, norm_idx)
            company_map[cand] = result
            stats["skip" if result["match_type"] == "__skip__" else "override"] += 1
            continue

        result = match_one(cand, dart_df, exact_idx, norm_idx, norm_names)
        company_map[cand] = result
        stats[result["match_type"]] += 1

        if result["match_type"] in ("fuzzy_review", "not_found"):
            alts = get_alts(normalize(cand), dart_df, norm_names)
            review_rows.append([
                cand, result["official_name"],
                round(result["confidence"] * 100),
                alts, "", "", "",
            ])

    total = len(candidates)
    auto  = stats["exact"] + stats["exact_normalized"] + stats["fuzzy_auto"] + stats["override"]
    need  = stats["fuzzy_review"] + stats["not_found"]

    print(f"\n[매칭 결과]  (고유 후보 {total:,}개)")
    print(f"  정확 일치          : {stats['exact']:,}")
    print(f"  정규화 일치        : {stats['exact_normalized']:,}")
    print(f"  퍼지 자동 확정     : {stats['fuzzy_auto']:,}")
    print(f"  퍼지 검토 필요     : {stats['fuzzy_review']:,}")
    print(f"  미발견 (검토 필요) : {stats['not_found']:,}")
    print(f"  수동 오버라이드    : {stats['override']:,}")
    print(f"  제외(SKIP)         : {stats['skip']:,}")
    print(f"\n  자동 처리 : {auto:,}건 ({auto/total*100:.1f}%)")
    print(f"  검토 필요 : {need:,}건 ({need/total*100:.1f}%)")

    print("\n저장 중...")
    save_company_master(company_map)
    save_project_with_company(projects, company_map)
    save_review_excel(review_rows)
    ensure_overrides_template()

    print("\n" + "=" * 60)
    if need:
        print("  [다음 단계]")
        print("  1. output/review_needed.xlsx 열기")
        print("  2. 'official_name_입력란' 채우기")
        print("  3. python import_review.py  → company_overrides.json 자동 반영")
        print("  4. python build_company_master.py 재실행")
    else:
        print("  모든 항목 자동 처리 완료!")
    print("=" * 60)


if __name__ == "__main__":
    main()
